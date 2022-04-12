import simplexalgo as SAS
import openpyxl as xl

wb = xl.load_workbook("H:\\MM485 - Operation Research\\test.xlsx")
ws = wb.active

#Assign a data type to the values we are reading 
Z, A, B, C = 0, [], [], []

#This loop assigns the Z row values from the excel
for column in range(2, ws.max_column):
    A.append(float(ws.cell(2,column).value))


#This assigns the B rows and the corresponding rhs from the excel 
for row in range(3, ws.max_row+1):

    B.append([])
    C.append(float(ws.cell(row, ws.max_column).value))

    for column in range(2, ws.max_column):

        B[row-3].append(float(ws.cell(row,column).value))

#This is the rhs of the Z row from the excel
Z = float(ws.cell(2, ws.max_column).value)

A,B,C,Z = SAS.s_m(Z,A,B,C, maximisation = True)

ws.cell(1, ws.max_column+3, "OPTIMIZED VALUE" )
ws.cell(1, ws.max_column+1, Z)

wb.save("H:\\MM485 - Operation Research\\test.xlsx")
